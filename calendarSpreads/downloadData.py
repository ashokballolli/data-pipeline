# -*- coding: utf-8 -*-
"""
This program fetches continous futures data for Index and stock futures in NSE
using NSEpy package
It will create an excel file with the current and near month futures Close and
Last traded price. To know more how to use it check the below link

https://zerodha.com/varsity/chapter/calendar-spreads/

@author: arun kamath
"""
######################################################
# Import section
import pandas as pd
from datetime import date, timedelta
from nsepy import get_history
from nsepy.derivatives import get_expiry_date
import datetime
from nsepy import get_quote
import os
import arrow
from openpyxl import load_workbook
import csv


def fetch_data(sym, expiry, start_date):
    df_next_temp = pd.DataFrame()
    df_curr_temp = pd.DataFrame()
    df_combined_temp = pd.DataFrame()

    if sym == "NIFTY" or sym == "BANKNIFTY":  # add OR cond. for additional index
        index_bool = True
    else:
        index_bool = False

    curr_fut_data = get_history(symbol=sym,
                                start=start_date,
                                end=expiry,
                                index=index_bool,
                                futures=True,
                                expiry_date=expiry)

    if expiry.month == 12:
        next_expiry = get_expiry_date(expiry.year + 1, 1)
    else:
        next_expiry = get_expiry_date(expiry.year, expiry.month + 1)
    # bug in nse site fetching wrong expiry for march, 2018
    if next_expiry == date(2018, 3, 29):
        next_expiry = date(2018, 3, 28)

    next_fut_data = get_history(symbol=sym,
                                start=start_date,
                                end=expiry,
                                index=index_bool,
                                futures=True,
                                expiry_date=next_expiry)

    df_curr_temp = curr_fut_data[["Last", "Close"]].copy()
    df_next_temp = next_fut_data[["Last", "Close"]].copy()

    df_curr_temp.reset_index(drop=False, inplace=True)
    df_curr_temp['Date'] = pd.to_datetime(df_curr_temp['Date'])
    df_curr_temp.set_index('Date', inplace=True)

    df_next_temp.reset_index(drop=False, inplace=True)
    df_next_temp['Date'] = pd.to_datetime(df_next_temp['Date'])
    df_next_temp.set_index('Date', inplace=True)

    # you can add/remove columns here
    df_combined_temp["Curr Close"] = df_curr_temp["Close"]
    # df_combined_temp["Curr Last"] = df_curr_temp["Last"]
    df_combined_temp["Next Close"] = df_next_temp["Close"]
    # df_combined_temp["Next Last"] = df_next_temp["Last"]

    return df_combined_temp


def get_data():
    # try to keep the number of contracts less than 8 for optimal performance
    symbols = ["ZEEL"]
    month = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]

    # add year at new year. Check in console whether old data is still fetched
    # if not fetched remove it and make suitable logic changes in line 95
    year = [2018, 2019]

    # if you have already run this program before,change the filename below to
    # avoid overwriting
    filename = "calendar_spread.xlsx"

    now_d = date.today()
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    for sym in symbols:
        df_combined = pd.DataFrame()
        for yy in year:
            for mm in month:
                expiry = get_expiry_date(year=yy, month=mm)

                if expiry == date(2018, 1, 25):
                    start_date = expiry + timedelta(days=1)

                else:
                    # bug in nse site resulting in wrong expiry for march, 2018
                    # can be removed if fixed by nse
                    # run get_expiry_date(2018, 3) in console to verify
                    if expiry == date(2018, 3, 29):
                        expiry = date(2018, 3, 28)
                    df_combined = df_combined.append(fetch_data(sym, expiry,
                                                                start_date))
                    start_date = expiry + timedelta(days=1)
                    if start_date > now_d:
                        df_combined.to_excel(writer, sym)
                        break
    writer.save()


def get_futures_data_continous(symbol, start_date, end_date):
    # try to keep the number of contracts less than 8 for optimal performance
    month = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    start_month = start_date.month
    end_month = end_date.month
    start_year = start_date.year
    end_year = end_date.year

    # add year at new year. Check in console whether old data is still fetched
    # if not fetched remove it and make suitable logic changes in line 95
    year = [2018, 2019]

    # if you have already run this program before,change the filename below to
    # avoid overwriting
    file = "../data/data_calendarSpreads/" + symbol + ".xlsx"
    exists = os.path.isfile(file)

    if exists:
        writerA = pd.ExcelWriter(file, engine='openpyxl')
        stock_data_df = pd.read_excel(file, sheet_name=symbol)
        last_record = stock_data_df['Date'].tail(1)
        last_record_date_time = last_record.iloc[0]
        last_record_date = last_record_date_time.to_pydatetime().date()
        current_date = datetime.datetime.now().date()
        current_year = current_date.year
        current_month = current_date.month

        if last_record_date < current_date - timedelta(days=1):
            expiry = get_expiry_date(year=current_year, month=current_month)
            new_data = fetch_data(symbol, expiry, last_record_date + timedelta(days=1))
            # stock_data_df.append(fetch_data(symbol, expiry, last_record_date + timedelta(days=1)))
            if (len(new_data.index) > 0):
                append_df_to_excel(file, new_data, sheet_name=symbol)


    else:
        writer = pd.ExcelWriter(file, engine='xlsxwriter')
        df_combined = pd.DataFrame()
        for yy in year:
            for mm in month:
                if ((mm >= start_month and yy == start_year) or yy > start_year):
                    expiry = get_expiry_date(year=yy, month=mm)

                    if expiry == date(2018, 1, 25):
                        start_date = expiry + timedelta(days=1)

                    else:
                        # bug in nse site resulting in wrong expiry for march, 2018
                        # can be removed if fixed by nse
                        # run get_expiry_date(2018, 3) in console to verify
                        if expiry == date(2018, 3, 29):
                            expiry = date(2018, 3, 28)
                        df_combined = df_combined.append(fetch_data(symbol, expiry,
                                                                    start_date))

                        start_date = expiry + timedelta(days=1)
                        if start_date > end_date:
                            df_combined.to_excel(writer, symbol)
                            break
        # writer.save()
        if (len(df_combined.index) > 200):
            writer.save()


def get_futures_data_LTP(symbol):
    current_date = datetime.datetime.now().date()
    current_year = current_date.year
    current_month = current_date.month
    current_expiry_date = get_expiry_date(year=current_year, month=current_month)
    if current_date > current_expiry_date:
        current_expiry_date = get_expiry_date(year=current_year, month=current_month+1)
        next_expiry_date = get_expiry_date(year=current_year, month=current_month + 2)
    else:
        next_expiry_date = get_expiry_date(year=current_year, month=current_month + 1)

    current_exp_quote = get_quote(symbol=symbol, instrument='FUTSTK', expiry=current_expiry_date, strike=450.00)
    next_exp_quote = get_quote(symbol=symbol, instrument='FUTSTK', expiry=next_expiry_date, strike=450.00)
    print(current_exp_quote)
    print(next_exp_quote)
    current_exp_quote_lastPrice = current_exp_quote['lastPrice']
    next_exp_quote_lastPrice = next_exp_quote['lastPrice']

    # make current_expiry_ltp = current_month_buyPrice1 from spread
    # make next_expiry_ltp = next_month_sellPrice1 from spread
    # becasue this will be only the sell spread: Buying current month, selling next month
    # current_exp_quote_lastPrice = current_exp_quote['sellPrice1']
    # next_exp_quote_lastPrice = next_exp_quote['buyPrice1']
    print(current_exp_quote_lastPrice)
    print(next_exp_quote_lastPrice)
    if current_exp_quote_lastPrice == '-':
        current_exp_quote_lastPrice = 0.0
    if next_exp_quote_lastPrice == '-':
        next_exp_quote_lastPrice = 0.0

    if current_exp_quote_lastPrice > 1 and next_exp_quote_lastPrice > 1:
        return current_exp_quote_lastPrice, next_exp_quote_lastPrice;
    else:
        return 0.0, 0.0;

def writeLTPFutureToDataFile(symbol):
    file = "../data/data_calendarSpreads/" + symbol + ".xlsx"
    stock_data_df = pd.read_excel(file, sheet_name=symbol)
    last_record = stock_data_df['Date'].tail(1)
    last_record_date_time = last_record.iloc[0]
    last_record_date = last_record_date_time.to_pydatetime().date()
    current_date_time = datetime.datetime.now()
    current_date = current_date_time.date()
    current_year = current_date.year
    current_month = current_date.month
    wb = load_workbook(filename=file)
    ws = wb.get_sheet_by_name(symbol)
    current_ltp, next_ltp = get_futures_data_LTP(symbol)
    if current_ltp==0.0 or next_ltp==0.0:
        print("******************** SOMETHING WRONG WITH THE LTP ********************")
    else:
        if last_record_date == current_date:
            action_row = ws.max_row
        else:
            action_row = ws.max_row + 1

        ws.cell(row=action_row, column=1).value = current_date_time
        ws.cell(row=action_row, column=2).value = current_ltp
        ws.cell(row=action_row, column=3).value = next_ltp
        wb.save(file)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, header=False, **to_excel_kwargs)

    # save the workbook
    writer.save()

def main():
    get_data()


if __name__ == '__main__':
    main()
